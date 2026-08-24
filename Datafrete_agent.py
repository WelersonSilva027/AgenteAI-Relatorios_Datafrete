"""
Agente de Automação - TMS Datafrete (FOTUS)
==========================================
Interface Gráfica com CustomTkinter e Relatórios com Logos FOTUS/DATAFRETE.

Dependências:
    pip install -r requirements.txt
    playwright install chromium

    (python-calamine e lxml são obrigatórios: sem eles, os arquivos exportados
    pelo Datafrete não conseguem ser lidos, pois o openpyxl sozinho não abre
    o estilo desses arquivos.)
"""

import os
import re
import asyncio
import zipfile
import threading
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from playwright.async_api import (
    async_playwright, Page, BrowserContext,
    TimeoutError as PlaywrightTimeout,
)
import customtkinter as ctk
import pandas as pd
import warnings

# Ignora avisos chatos do Excel no terminal
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

load_dotenv()

# ── Configurações gerais ───────────────────────────────────────────────────────
# A URL base do Datafrete pode vir do .env, mas o login agora é exigido na interface
BASE_URL = os.getenv("DATAFRETE_URL") or "https://tms.datafrete.com.br"
DOWNLOAD_DIR = Path("./exports")
HEADLESS     = False

# ── Sistema de Log para a UI ───────────────────────────────────────────────────
UI_LOG_CALLBACK = None

def log(msg: str):
    """Loga no terminal e também envia para a Interface Gráfica, se conectada."""
    timestamp = datetime.now().strftime('%H:%M:%S')
    format_msg = f"[{timestamp}] {msg}"
    print(format_msg)
    if UI_LOG_CALLBACK:
        UI_LOG_CALLBACK(format_msg)


# ── XPaths por tipo de card ────────────────────────────────────────────────────
CARDS = {
    "arquivos": {
        "label":         "Pendências de Arquivos",
        "prefixo":       "Pendencias_Arquivos",
        "prefixo_bruto": "_bruto_arquivos",
        "xpath_lupa":    (
            "xpath=/html/body/div[5]/div/div[1]/div[1]/div/div/div[2]"
            "/div[1]/div/div/div[3]/div[2]/button"
        ),
        "radio_for":     "radio-empresa-transportadorpendencia-arquivo",
        "xpath_exportar":(
            "xpath=/html/body/div[5]/div/div[1]/div[1]/div/div/div[2]"
            "/div[2]/div/div/div[3]/div/button[2]"
        ),
        "col_a_resumo":  "Quantidade NF sem CT",
        "col_b_resumo":  "Quantidade CT sem NF",
        "col_a_label":   "NF sem CT",
        "col_b_label":   "CT sem NF",
        "cor_banner_a":  "1F6B3A",   # verde
        "cor_banner_b":  "4A235A",   # roxo
        "titulo_aba_a":  "NF sem CT",
        "titulo_aba_b":  "CT sem NF",
    },
    "entrega": {
        "label":         "Pendências de Entrega",
        "prefixo":       "Pendencias_Entrega",
        "prefixo_bruto": "_bruto_entrega",
        "xpath_lupa":    (
            "xpath=/html/body/div[5]/div/div[1]/div[1]/div/div/div[5]"
            "/div[1]/div/div/div[3]/div[2]/button"
        ),
        "radio_xpath":   (
            "xpath=/html/body/div[5]/div/div[1]/div[1]/div/div/div[5]"
            "/div[2]/div/div/div[2]/div[1]/div[2]/label"
        ),
        "xpath_exportar":(
            "xpath=/html/body/div[5]/div/div[1]/div[1]/div/div/div[5]"
            "/div[2]/div/div/div[3]/div/button[2]"
        ),
        "col_a_resumo":  "Quantidade Pendente",
        "col_b_resumo":  "Quantidade Ocorrência",
        "col_a_label":   "Pendente",
        "col_b_label":   "Ocorrência",
        "cor_banner_a":  "7B3F00",   # marrom
        "cor_banner_b":  "1A3A5C",   # azul escuro
        "titulo_aba_a":  "Pendente",
        "titulo_aba_b":  "Ocorrência",
    },
}

XPATHS_OPCOES = [
    "xpath=/html/body/div[6]/div[1]/div[6]/div[2]/div/button[2]/span",
    "xpath=/html/body/div[5]/div[1]/div[6]/div[2]/div/button[2]/span",
    "xpath=/html/body/div[6]/div[1]/div[5]/div[2]/div/button[2]/span",
    "xpath=/html/body/div[5]/div[1]/div[5]/div[2]/div/button[2]/span",
    "button:has-text('Opções')",
]
XPATHS_EXPORTAR_LISTAGEM = [
    "xpath=/html/body/div[6]/div[1]/div[6]/div[2]/div/div[2]/ul/li[4]/a",
    "xpath=/html/body/div[5]/div[1]/div[6]/div[2]/div/div[2]/ul/li[4]/a",
    "xpath=/html/body/div[6]/div[1]/div[5]/div[2]/div/div[2]/ul/li[4]/a",
    "xpath=/html/body/div[5]/div[1]/div[5]/div[2]/div/div[2]/ul/li[4]/a",
    "a:has-text('Exportar listagem')",
    "li:has-text('Exportar listagem') a",
]

# Paleta
AZUL_FOTUS  = "1B4F8A"
AZUL_CLARO  = "D6E4F7"
AMARELO     = "FFC000"
CINZA_LINHA = "F2F2F2"
BRANCO      = "FFFFFF"


def sanitizar_nome(texto: str) -> str:
    texto = re.sub(r'[\\/:*?"<>|]', '-', texto)
    texto = re.sub(r'\s+', '_', texto)
    return re.sub(r'_+', '_', texto).strip('_')


def abrir_arquivo(caminho: Path):
    """Abre o Excel gerado no programa padrão do Windows (normalmente o Excel)."""
    try:
        os.startfile(str(caminho))
        log(f"  🗂  Abrindo relatório: {caminho.name}")
    except Exception as e:
        log(f"  ⚠ Não foi possível abrir '{caminho.name}' automaticamente: {e}")


def nome_aba(unidade: str, sufixo: str) -> str:
    completo = f"{unidade} - {sufixo}"
    if len(completo) <= 31:
        return completo
    espaco = 31 - len(sufixo) - 3
    return f"{unidade[:espaco]} - {sufixo}"


# ══════════════════════════════════════════════════════════════════════════════
# CACHE
# ══════════════════════════════════════════════════════════════════════════════

def verificar_cache(card: str) -> tuple[dict, bool]:
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    prefixo = CARDS[card]["prefixo_bruto"].replace("_bruto_", "_detalhe_")
    arquivos = sorted(DOWNLOAD_DIR.glob(f"_detalhe_{card}_*.xlsx"))
    if not arquivos:
        return {}, False
    detalhes_cache = defaultdict(list)
    for arq in arquivos:
        stem = arq.stem.replace(f"_detalhe_{card}_", "")
        for tipo in ("NF_sem_CT", "CT_sem_NF", "Pendente", "Ocorrencia"):
            if stem.endswith(tipo):
                resto  = stem[:-(len(tipo) + 1)]
                partes = resto.split("_")
                unidade = f"{partes[0]} {partes[1]}" if len(partes) >= 2 else partes[0]
                transp  = " ".join(partes[2:]).replace("_", " ")
                detalhes_cache[unidade].append(
                    {"arquivo": arq, "tipo": tipo, "transp": transp}
                )
                break
    return detalhes_cache, bool(detalhes_cache)


# ══════════════════════════════════════════════════════════════════════════════
# PLAYWRIGHT
# ══════════════════════════════════════════════════════════════════════════════

async def step_login(page: Page, email: str, senha: str):
    log("ETAPA 1 — Login")
    await page.goto(BASE_URL, wait_until="networkidle")
    for sel in ["input[type='email']", "input[name*='email']", "input[name*='usuario']"]:
        try:
            await page.fill(sel, email, timeout=3000)
            log(f"  → Usuário preenchido ({sel})")
            break
        except Exception:
            continue
    await page.fill("input[type='password']", senha)
    for sel in ["button:has-text('Continuar')", "button[type='submit']", "button:has-text('Entrar')"]:
        try:
            await page.click(sel, timeout=3000)
            break
        except Exception:
            continue
    else:
        await page.press("input[type='password']", "Enter")
    await page.wait_for_load_state("networkidle")
    log("  ✓ Login concluído")


async def step_fechar_modais(page: Page):
    try:
        fechar = page.locator("button:has-text('Fechar')")
        if await fechar.is_visible(timeout=1500):
            await fechar.click()
            await asyncio.sleep(0.8)
    except Exception:
        pass
    await page.keyboard.press("Escape")
    await asyncio.sleep(0.4)


async def step_abrir_modal(page: Page, cfg: dict):
    log(f"ETAPA 2 — Abrindo modal: {cfg['label']}")
    await page.wait_for_load_state("networkidle")
    await asyncio.sleep(1)
    await step_fechar_modais(page)

    lupa = page.locator(cfg["xpath_lupa"])
    await lupa.wait_for(state="visible", timeout=10000)
    await lupa.click()
    await asyncio.sleep(1)

    if "radio_for" in cfg:
        await page.wait_for_selector(
            f"label[for='{cfg['radio_for']}']", timeout=10000
        )
    else:
        await page.wait_for_selector(cfg["radio_xpath"], timeout=10000)
    log("  ✓ Modal aberto")


async def step_selecionar_visualizacao(page: Page, cfg: dict):
    log("ETAPA 3 — Selecionando visualização Organização + Transportador")
    await asyncio.sleep(0.5)

    if "radio_for" in cfg:
        label = page.locator(f"label[for='{cfg['radio_for']}']")
    else:
        label = page.locator(cfg["radio_xpath"])

    await label.wait_for(state="visible", timeout=5000)
    await label.click()
    await asyncio.sleep(1.5)
    log("  ✓ Visualização selecionada")


async def step_exportar_resumo(page: Page, cfg: dict) -> Path:
    log("ETAPA 4 — Exportando resumo")
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")

    btn = page.locator(cfg["xpath_exportar"])
    await btn.wait_for(state="visible", timeout=10000)
    await asyncio.sleep(0.5)

    async with page.expect_download(timeout=30000) as dl_info:
        await btn.click()

    download = await dl_info.value
    ext      = Path(download.suggested_filename).suffix or ".xlsx"
    filepath = DOWNLOAD_DIR / f"{cfg['prefixo_bruto']}_{timestamp}{ext}"
    await download.save_as(filepath)
    log(f"  ✓ Resumo salvo: {filepath.name}")
    return filepath


async def _clicar_primeiro_que_funcionar(page: Page, xpaths: list, timeout=8000):
    for xpath in xpaths:
        try:
            loc = page.locator(xpath)
            await loc.wait_for(state="visible", timeout=timeout)
            await loc.click()
            return xpath
        except Exception:
            continue
    raise RuntimeError("Nenhum seletor funcionou.")


async def _exportar_listagem_na_aba(nova_aba: Page, label: str) -> Path | None:
    try:
        await nova_aba.wait_for_load_state("networkidle", timeout=20000)
        await asyncio.sleep(1.5)
        await _clicar_primeiro_que_funcionar(nova_aba, XPATHS_OPCOES, timeout=12000)
        await asyncio.sleep(0.8)
        async with nova_aba.expect_download(timeout=30000) as dl_info:
            await _clicar_primeiro_que_funcionar(
                nova_aba, XPATHS_EXPORTAR_LISTAGEM, timeout=6000
            )
        download = await dl_info.value
        ext      = Path(download.suggested_filename).suffix or ".xlsx"
        nome     = sanitizar_nome(label)[:80]
        filepath = DOWNLOAD_DIR / f"_detalhe_{nome}{ext}"
        await download.save_as(filepath)
        log(f"    ✓ Listagem salva: {filepath.name}")
        return filepath
    except Exception as e:
        log(f"    ⚠ Falha ao exportar '{label}': {e}")
        return None


async def step_drill_down(page: Page, context: BrowserContext, card: str, cfg: dict) -> dict:
    log("ETAPA 5 — Drill-down por unidade/transportador")
    detalhes = defaultdict(list)
    await asyncio.sleep(1)

    linhas_locator = page.locator(
        "div[role='dialog'] table tbody tr, "
        "div.modal table tbody tr, "
        ".modal-content table tbody tr"
    )
    n_linhas = await linhas_locator.count()
    if n_linhas == 0:
        linhas_locator = page.locator("table tbody tr")
        n_linhas = await linhas_locator.count()

    log(f"  → {n_linhas} linha(s) na tabela do modal")
    if n_linhas == 0:
        log("  ⚠ Tabela não localizada. Pulando drill-down.")
        return detalhes

    celulas = []
    for i in range(n_linhas):
        linha = linhas_locator.nth(i)
        cols  = linha.locator("td")
        if await cols.count() < 4:
            continue
        org_raw = (await cols.nth(0).inner_text()).strip()
        transp  = (await cols.nth(1).inner_text()).strip()
        if org_raw.lower() in ("total", "total geral", "totais"):
            continue
        nome = org_raw.split("(")[0].strip() if "(" in org_raw else org_raw
        try:    val_a = int((await cols.nth(2).inner_text()).strip())
        except: val_a = 0
        try:    val_b = int((await cols.nth(3).inner_text()).strip())
        except: val_b = 0

        tipo_a = "NF_sem_CT"  if card == "arquivos" else "Pendente"
        tipo_b = "CT_sem_NF"  if card == "arquivos" else "Ocorrencia"

        if val_a > 0:
            celulas.append({"unidade": nome, "transp": transp,
                            "tipo": tipo_a, "valor": val_a, "row": i, "col": 2})
        if val_b > 0:
            celulas.append({"unidade": nome, "transp": transp,
                            "tipo": tipo_b, "valor": val_b, "row": i, "col": 3})

    log(f"  → {len(celulas)} célula(s) para exportar")

    for info in celulas:
        unidade = info["unidade"]
        transp  = info["transp"]
        tipo    = info["tipo"]
        label   = f"{card}_{unidade}_{transp}_{tipo}"
        log(f"  → {unidade} | {transp} | {tipo} = {info['valor']}")
        try:
            await asyncio.sleep(0.5)
            celula_loc = linhas_locator.nth(info["row"]).locator("td").nth(info["col"])
            async with context.expect_page(timeout=15000) as nova_aba_info:
                await celula_loc.click()
            nova_aba = await nova_aba_info.value
            await nova_aba.bring_to_front()
            filepath = await _exportar_listagem_na_aba(nova_aba, label)
            if filepath:
                detalhes[unidade].append(
                    {"arquivo": filepath, "tipo": tipo, "transp": transp}
                )
            await nova_aba.close()
            await page.bring_to_front()
            await asyncio.sleep(0.6)
        except PlaywrightTimeout:
            log(f"    ⚠ Timeout para '{label}'")
            try:   await nova_aba.close()
            except: pass
            await page.bring_to_front()
        except Exception as e:
            log(f"    ⚠ Erro em '{label}': {e}")
            await page.bring_to_front()

    log(f"  ✓ Drill-down concluído — {sum(len(v) for v in detalhes.values())} arquivo(s)")
    return detalhes


# ══════════════════════════════════════════════════════════════════════════════
# LEITURA DOS ARQUIVOS (MUITO MAIS ROBUSTO)
# ══════════════════════════════════════════════════════════════════════════════

def _normalizar_strings(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converte todas as colunas para dtype "object" com strings Python puras.
    Necessário porque motores como calamine/pandas mais novos podem devolver
    colunas em dtypes de extensão (ex.: "str" nativo do pandas), cujo acesso
    célula-a-célula em certos casos (.iloc, .values) pode devolver um Series
    em vez de um escalar — o que gera lixo tipo "Name: 3, dtype: str" nas células.
    """
    df = df.copy()
    df.columns = [str(c) for c in df.columns]
    for col in df.columns:
        df[col] = df[col].astype(object).where(df[col].notna(), "")
    return df


def _ler_xlsx(path: Path):
    """
    Tenta abrir o Excel gerado pelo Datafrete utilizando múltiplos motores,
    para garantir que os textos não venham ocultos ou vazios.
    """
    # 1. Tentativa primária: calamine (motor mais tolerante a exports não-Microsoft,
    #    como os do Datafrete — o openpyxl costuma rejeitar o estilo desses arquivos)
    try:
        df = pd.read_excel(path, engine="calamine")
        if df is not None and not df.empty and len(df.columns) > 1:
            return _normalizar_strings(df)
    except Exception as e:
        log(f"  ⚠ Falha ao ler '{path.name}' via calamine: {e}")

    # 2. Tentativa secundária: openpyxl
    try:
        df = pd.read_excel(path, engine="openpyxl")
        if df is not None and not df.empty and len(df.columns) > 1:
            return _normalizar_strings(df)
    except Exception as e:
        log(f"  ⚠ Falha ao ler '{path.name}' via openpyxl: {e}")

    # 3. Tentativa terciária: Ler como tabela HTML (muitos sistemas usam .xls falso)
    try:
        dfs = pd.read_html(path, decimal=',', thousands='.')
        if dfs:
            return _normalizar_strings(dfs[0])
    except Exception as e:
        log(f"  ⚠ Falha ao ler '{path.name}' via read_html: {e}")

    # 4. Fallback extremo via XML parsing (se der problema severo nas libs)
    try:
        NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(path) as z:
            sheet  = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet")][0]
            tree   = ET.parse(z.open(sheet))
            shared = []
            if "xl/sharedStrings.xml" in z.namelist():
                ss = ET.parse(z.open("xl/sharedStrings.xml"))
                shared = [
                    si.text or "".join(t.text or "" for t in si.iter(f"{{{NS}}}t"))
                    for si in ss.findall(f".//{{{NS}}}si")
                ]
            rows_data = []
            for row in tree.findall(f".//{{{NS}}}row"):
                vals = []
                for cell in row.findall(f"{{{NS}}}c"):
                    v = cell.find(f"{{{NS}}}v")
                    t = cell.get("t", "")
                    if v is None:   
                        vals.append("")
                    elif t == "s":  
                        try: vals.append(shared[int(v.text)])
                        except: vals.append(v.text)
                    else:           
                        vals.append(v.text)
                rows_data.append(vals)
        if rows_data:
            n_cols = len(rows_data[0])
            corpo  = [
                (linha + [""] * (n_cols - len(linha)))[:n_cols]
                for linha in rows_data[1:]
            ]
            return _normalizar_strings(pd.DataFrame(corpo, columns=rows_data[0]))
    except Exception as e:
        log(f"  ⚠ Falha ao ler '{path.name}' via XML manual: {e}")

    return None


def _ler_xlsx_com_retry(path: Path, tentativas: int = 4, espera: float = 1.5):
    """
    Reexecuta a leitura algumas vezes com espera entre elas.
    Necessário porque o arquivo acabou de ser baixado dentro de uma pasta
    sincronizada pelo OneDrive: por uma fração de segundo o OneDrive pode
    manter um lock exclusivo no arquivo recém-criado, o que faz calamine/
    openpyxl falharem mesmo com o arquivo perfeitamente válido.
    """
    import time
    for tentativa in range(1, tentativas + 1):
        df = _ler_xlsx(path)
        if df is not None and not df.empty:
            return df
        if tentativa < tentativas:
            log(f"  ⚠ Leitura vazia/falhou (tentativa {tentativa}/{tentativas}) — "
                f"aguardando arquivo ser liberado (OneDrive?) e tentando de novo...")
            time.sleep(espera)
    return None


def ler_arquivo_bruto(arquivo_bruto: Path):
    log(f"  → Lendo bruto: {arquivo_bruto.name}")
    df = _ler_xlsx_com_retry(arquivo_bruto)

    if df is None or df.empty:
        raise RuntimeError(f"Não foi possível ler os dados do arquivo {arquivo_bruto.name}")

    # Limpeza avançada: se as primeiras linhas vierem vazias, busca o cabeçalho real
    if all("unnamed" in str(c).lower() or str(c).strip() == "" for c in df.columns[:2]):
        for i, row in df.head(10).iterrows():
            if any("organiza" in str(v).lower() for v in row.values):
                df.columns = [str(v) for v in row.values]
                df = df.iloc[i+1:].reset_index(drop=True)
                break

    def acha_idx(termo):
        for idx, c in enumerate(df.columns):
            if termo.lower() in str(c).lower():
                return idx
        return None

    def primeiro(*valores, default):
        for v in valores:
            if v is not None:
                return v
        return default

    # NOVO: Puxando colunas unicamente pelas Posições (Índices) ao invés do nome
    # Assim, não importa se o nome vier vazio, duplicado ou com espaçamento.
    idx_org    = primeiro(acha_idx("organiza"), default=0)
    idx_transp = primeiro(acha_idx("transportador"), default=1)
    idx_nfct   = primeiro(
        acha_idx("nf sem ct"), acha_idx("nf_sem_ct"), acha_idx("pendente"), acha_idx("ct"),
        default=2,
    )
    idx_ctnf   = primeiro(
        acha_idx("ct sem nf"), acha_idx("ct_sem_nf"), acha_idx("ocorr"),
        default=3 if len(df.columns) > 3 else 2,
    )

    log(f"  → Mapeamento de Posição: ORG=[{idx_org}] TRANSP=[{idx_transp}] NF=[{idx_nfct}] CT=[{idx_ctnf}]")

    resumo      = defaultdict(lambda: [0, 0])
    por_unidade = {}

    for _, row in df.iterrows():
        try:
            # .iat garante um escalar (nunca um Series) na posição exata da célula
            org_raw = str(row.iat[idx_org]).strip()
        except Exception:
            continue

        # Pula as linhas vazias ou as que são apenas totais gerais no fim da tabela
        if not org_raw or org_raw.lower() in ("nan", "none", "total", "total geral", "totais"):
            continue

        nome = org_raw.split("(")[0].strip() if "(" in org_raw else org_raw

        try:
            nf = int(float(str(row.iat[idx_nfct]).replace(",", ".")))
        except:
            nf = 0

        try:
            ct = int(float(str(row.iat[idx_ctnf]).replace(",", "."))) if idx_ctnf < len(row) else 0
        except:
            ct = 0

        resumo[nome][0] += nf
        resumo[nome][1] += ct

        try:
            transp = str(row.iat[idx_transp]).strip()
            if not transp or transp.lower() in ("nan", "none", ""):
                transp = "Sem transportador"
        except:
            transp = "Sem transportador"
            
        por_unidade.setdefault(nome, {}).setdefault(transp, [0, 0])
        por_unidade[nome][transp][0] += nf
        por_unidade[nome][transp][1] += ct

    linhas   = sorted(resumo.items(), key=lambda x: x[1][0], reverse=True)
    total_nf = sum(v[0] for _, v in linhas)
    total_ct = sum(v[1] for _, v in linhas)

    log(f"  → Extraído: {len(linhas)} unidades | Col A (Soma): {total_nf} | Col B (Soma): {total_ct}")

    if not linhas:
        raise RuntimeError(
            f"Nenhuma unidade foi extraída de '{arquivo_bruto.name}'. "
            f"O arquivo abriu, mas nenhuma linha reconhecível de dados foi encontrada "
            f"(colunas detectadas: {list(df.columns)}). Tente executar novamente."
        )

    return linhas, total_nf, total_ct, por_unidade


def _bruto_mais_recente(card: str) -> Path | None:
    prefixo = CARDS[card]["prefixo_bruto"]
    candidatos = sorted(DOWNLOAD_DIR.glob(f"{prefixo}_*.xlsx"), reverse=True)
    return candidatos[0] if candidatos else None


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE ESTILO E GERAÇÃO DE EXCEL COM LOGOS
# ══════════════════════════════════════════════════════════════════════════════

def _make_styles():
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin  = Side(style="thin",   color="CCCCCC")
    med   = Side(style="medium", color=AZUL_FOTUS)
    B_TAB = Border(left=thin, right=thin, top=thin,  bottom=thin)
    B_HDR = Border(left=med,  right=med,  top=med,   bottom=med)

    def cs(ws, row, col, value="", bold=False, size=11,
            bg=None, fg="000000", align="center", border=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, size=size, color=fg, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center")
        if bg:      c.fill   = PatternFill("solid", fgColor=bg)
        if border: c.border = border
        return c

    return cs, B_TAB, B_HDR


def _banner(ws, texto, n_colunas, cor_fundo, subtitulo,
           logo_esq: str = None, logo_dir: str = None, linha_unidade: str = None) -> int:
    """
    Monta o cabeçalho da aba. Se `linha_unidade` for informado, o título fica
    curto (ex.: "Pendências de Arquivos") na linha 1 — junto com as logos — e o
    nome da unidade (ex.: "FOTUS ES") ganha sua PRÓPRIA linha, sem logos, para
    nunca disputar espaço/sobrepor as imagens. Retorna a primeira linha livre
    após o cabeçalho (onde o chamador deve começar a desenhar a tabela).
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # ── Linha 1: fundo azul + título (curto, para nunca colidir com as logos) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_colunas)
    c = ws.cell(row=1, column=1, value=texto)
    c.font      = Font(bold=True, size=16, color=BRANCO, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=cor_fundo)
    ws.row_dimensions[1].height = 40 if linha_unidade else 50

    # Pinta todas as células da faixa
    for col in range(1, n_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor=cor_fundo)

    pasta = Path(__file__).parent

    def _inserir_logo(nome_arquivo, ancora, altura_desejada=40):
        caminho = pasta / nome_arquivo
        if caminho.exists():
            try:
                img = XLImage(str(caminho))
                proporcao = altura_desejada / img.height
                img.width = int(img.width * proporcao)
                img.height = altura_desejada
                img.anchor = ancora
                ws.add_image(img)
            except Exception as e:
                log(f"Aviso: Não foi possível carregar a logo {nome_arquivo}: {e}")

    if logo_esq:
        _inserir_logo(logo_esq, "A1", altura_desejada=45)

    if logo_dir:
        coluna_final = get_column_letter(n_colunas)
        _inserir_logo(logo_dir, f"{coluna_final}1", altura_desejada=40)

    linha_seguinte = 2

    # ── Linha extra opcional: nome da unidade, em linha própria (sem logos) ──
    if linha_unidade:
        ws.merge_cells(start_row=linha_seguinte, start_column=1,
                       end_row=linha_seguinte, end_column=n_colunas)
        cu = ws.cell(row=linha_seguinte, column=1, value=linha_unidade)
        cu.font      = Font(bold=True, size=13, color=BRANCO, name="Calibri")
        cu.alignment = Alignment(horizontal="center", vertical="center")
        cu.fill      = PatternFill("solid", fgColor=cor_fundo)
        ws.row_dimensions[linha_seguinte].height = 24
        linha_seguinte += 1

    # ── Linha seguinte: faixa amarela + subtítulo ──
    ws.merge_cells(start_row=linha_seguinte, start_column=1,
                   end_row=linha_seguinte, end_column=n_colunas)
    c2 = ws.cell(row=linha_seguinte, column=1, value=subtitulo)
    c2.font      = Font(bold=True, size=11, color="1B4F8A", name="Calibri")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill      = PatternFill("solid", fgColor=AMARELO)
    ws.row_dimensions[linha_seguinte].height = 20

    return linha_seguinte + 1


def _montar_aba_resumo(wb, ws, cfg, titulo, dados, tot_a, tot_b, col_a_header="Unidade", unidade=None):
    cs, B_TAB, B_HDR = _make_styles()
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26

    linha_cab = _banner(
        ws=ws,
        texto=titulo,
        n_colunas=3,
        cor_fundo=AZUL_FOTUS,
        subtitulo=f"RELATÓRIO MATUTINO  —  {datetime.now().strftime('%d/%m/%Y')}",
        logo_esq="logo_fotus_excel.png",
        logo_dir="logo_datafrete_excel.png",
        linha_unidade=unidade,
    )

    ws.row_dimensions[linha_cab].height = 18
    cs(ws, linha_cab, 1, col_a_header,        bold=True, size=11, bg=AZUL_FOTUS, fg=BRANCO, border=B_HDR)
    cs(ws, linha_cab, 2, cfg["col_a_resumo"], bold=True, size=11, bg=AZUL_FOTUS, fg=BRANCO, border=B_HDR)
    cs(ws, linha_cab, 3, cfg["col_b_resumo"], bold=True, size=11, bg=AZUL_FOTUS, fg=BRANCO, border=B_HDR)

    for i, (nome, (va, vb)) in enumerate(dados):
        r   = linha_cab + 1 + i
        bg = CINZA_LINHA if i % 2 == 0 else BRANCO
        ws.row_dimensions[r].height = 18
        cs(ws, r, 1, nome, bg=bg, align="left",   border=B_TAB)
        cs(ws, r, 2, va,   bg=bg, align="center", border=B_TAB)
        cs(ws, r, 3, vb,   bg=bg, align="center", border=B_TAB)

    tr = linha_cab + 1 + len(dados)
    ws.row_dimensions[tr].height = 20
    cs(ws, tr, 1, "Total Geral", bold=True, size=11, bg=AZUL_CLARO, fg=AZUL_FOTUS, align="left",   border=B_HDR)
    cs(ws, tr, 2, tot_a,         bold=True, size=11, bg=AZUL_CLARO, fg=AZUL_FOTUS, align="center", border=B_HDR)
    cs(ws, tr, 3, tot_b,         bold=True, size=11, bg=AZUL_CLARO, fg=AZUL_FOTUS, align="center", border=B_HDR)


def _montar_aba_detalhes(wb, titulo_aba, titulo_banner, cor_banner,
                         subtitulo, itens_detalhes, cs, B_TAB, B_HDR, unidade=None):
    """
    Monta uma listagem ÚNICA e contínua com todos os transportadores juntos
    (a própria coluna "Transportador" já vem no arquivo exportado pelo
    Datafrete, então dá para filtrar por ela sem precisar separar em blocos).
    """
    ws = wb.create_sheet(title=titulo_aba)

    colunas     = None
    todas_linhas = []
    for info in itens_detalhes:
        df_det = _ler_xlsx(info["arquivo"])
        if df_det is None or df_det.empty:
            continue
            
        # Manter apenas as colunas de A até L (as 12 primeiras colunas)
        df_det = df_det.iloc[:, :12]

        if colunas is None:
            colunas = list(df_det.columns)
        else:
            df_det = df_det.reindex(columns=colunas, fill_value="")
        todas_linhas.extend(df_det.itertuples(index=False, name=None))

    if colunas is None:
        colunas = ["Sem dados"]
    n_cols = len(colunas)

    for col_i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 18

    linha_cab = 1

    ws.row_dimensions[linha_cab].height = 18
    for col_i, col_name in enumerate(colunas, start=1):
        cs(ws, linha_cab, col_i, str(col_name),
           bold=True, size=9, bg="2E6DB4", fg=BRANCO, align="center", border=B_TAB)
        ws.column_dimensions[get_column_letter(col_i)].width = max(
            ws.column_dimensions[get_column_letter(col_i)].width or 8,
            min(len(str(col_name)) + 4, 40))

    linha_atual = linha_cab + 1
    for i_row, valores in enumerate(todas_linhas):
        ws.row_dimensions[linha_atual].height = 15
        bg = CINZA_LINHA if i_row % 2 == 0 else BRANCO
        for col_i, val in enumerate(valores, start=1):
            cs(ws, linha_atual, col_i, val,
               size=9, bg=bg, align="left", border=B_TAB)
            ws.column_dimensions[get_column_letter(col_i)].width = max(
                ws.column_dimensions[get_column_letter(col_i)].width or 8,
                min(len(str(val)) + 2, 40))
        linha_atual += 1

    if todas_linhas:
        ultima_coluna = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A{linha_cab}:{ultima_coluna}{linha_atual - 1}"

    log(f"    → Aba '{titulo_aba}': {len(todas_linhas)} linha(s)")
    return ws


def gerar_excel_geral(linhas, total_nf, total_ct, por_unidade, cfg, card) -> Path:
    import openpyxl
    log(f"ETAPA 5 — Gerando Excel Geral ({cfg['label']})")
    cs, B_TAB, B_HDR = _make_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Geral"

    _montar_aba_resumo(wb, ws, cfg,
                       f"FOTUS  ·  {cfg['label'].upper()}  ·  DATAFRETE",
                       linhas, total_nf, total_ct, col_a_header="Unidade Fotus")

    for unidade, transportadores in sorted(por_unidade.items()):
        ws_u     = wb.create_sheet(title=nome_aba(unidade, "Resumo"))
        linhas_u = sorted(transportadores.items(), key=lambda x: x[1][0], reverse=True)
        _montar_aba_resumo(wb, ws_u, cfg,
                           cfg["label"],
                           linhas_u,
                           sum(v[0] for _, v in linhas_u),
                           sum(v[1] for _, v in linhas_u),
                           col_a_header="Transportador",
                           unidade=unidade)

    timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    saida     = DOWNLOAD_DIR / f"{cfg['prefixo']}_Geral_{timestamp}.xlsx"
    wb.save(saida)
    log(f"  ✓ Excel salvo: {saida.name}")
    return saida


def gerar_excel_detalhado(linhas, total_nf, total_ct, por_unidade, detalhes, cfg, card) -> Path:
    import openpyxl
    log(f"ETAPA 6 — Gerando Excel Detalhado ({cfg['label']})")
    cs, B_TAB, B_HDR = _make_styles()

    # ── CRIANDO PASTA DIÁRIA ──
    data_hoje_pasta = datetime.now().strftime("%Y-%m-%d")
    pasta_importacao = DOWNLOAD_DIR / f"importacao_{data_hoje_pasta}"
    pasta_importacao.mkdir(exist_ok=True)
    log(f"  → Pasta de importação: {pasta_importacao.name}")

    wb        = openpyxl.Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Geral"
    data_hoje = datetime.now().strftime("%d/%m/%Y")

    _montar_aba_resumo(wb, ws_resumo, cfg,
                       f"FOTUS  ·  {cfg['label'].upper()}  ·  DATAFRETE",
                       linhas, total_nf, total_ct, col_a_header="Unidade Fotus")
    log(f"  → Aba 'Resumo Geral' ({len(linhas)} unidades)")

    todas_unidades = sorted(set(list(por_unidade.keys()) + list(detalhes.keys())))

    for unidade in todas_unidades:
        log(f"  → Processando: {unidade}")

        transportadores = por_unidade.get(unidade, {})
        linhas_u = sorted(transportadores.items(), key=lambda x: x[1][0], reverse=True)
        ws_u = wb.create_sheet(title=nome_aba(unidade, "Resumo"))
        _montar_aba_resumo(wb, ws_u, cfg,
                           cfg["label"],
                           linhas_u,
                           sum(v[0] for _, v in linhas_u),
                           sum(v[1] for _, v in linhas_u),
                           col_a_header="Transportador",
                           unidade=unidade)

        def salvar_aba_individual(itens, tipo_aba, nome_aba_original, label_coluna, cor_banner, label_resumo):
            # 1. Adiciona no arquivão
            _montar_aba_detalhes(
                wb,
                titulo_aba     = nome_aba_original,
                titulo_banner  = label_coluna,
                cor_banner     = cor_banner,
                subtitulo      = f"{label_resumo.upper()}  —  {data_hoje}",
                itens_detalhes = itens,
                cs=cs, B_TAB=B_TAB, B_HDR=B_HDR,
                unidade=unidade,
            )
            
            # 2. Cria arquivo isolado
            wb_indiv = openpyxl.Workbook()
            _montar_aba_detalhes(
                wb_indiv,
                titulo_aba     = "Detalhes",
                titulo_banner  = label_coluna,
                cor_banner     = cor_banner,
                subtitulo      = f"{label_resumo.upper()}  —  {data_hoje}",
                itens_detalhes = itens,
                cs=cs, B_TAB=B_TAB, B_HDR=B_HDR,
                unidade=unidade,
            )
            # Limpa aba vazia padrão do openpyxl
            if "Sheet" in wb_indiv.sheetnames and len(wb_indiv.sheetnames) > 1:
                wb_indiv.remove(wb_indiv["Sheet"])
            
            safe_unidade = re.sub(r'[\\/:*?"<>|]', '-', unidade)
            nome_arquivo = f"{safe_unidade} - {tipo_aba}.xlsx"
            caminho_indiv = pasta_importacao / sanitizar_nome(nome_arquivo).replace('_-_', ' - ')
            wb_indiv.save(caminho_indiv)

        tipo_a  = "NF_sem_CT" if card == "arquivos" else "Pendente"
        itens_a = [i for i in detalhes.get(unidade, []) if i["tipo"] == tipo_a]
        if itens_a:
            salvar_aba_individual(
                itens=itens_a,
                tipo_aba=cfg["titulo_aba_a"],
                nome_aba_original=nome_aba(unidade, cfg["titulo_aba_a"]),
                label_coluna=cfg["col_a_label"],
                cor_banner=cfg["cor_banner_a"],
                label_resumo=cfg["col_a_resumo"]
            )

        tipo_b  = "CT_sem_NF" if card == "arquivos" else "Ocorrencia"
        itens_b = [i for i in detalhes.get(unidade, []) if i["tipo"] == tipo_b]
        if itens_b:
            salvar_aba_individual(
                itens=itens_b,
                tipo_aba=cfg["titulo_aba_b"],
                nome_aba_original=nome_aba(unidade, cfg["titulo_aba_b"]),
                label_coluna=cfg["col_b_label"],
                cor_banner=cfg["cor_banner_b"],
                label_resumo=cfg["col_b_resumo"]
            )

    timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    saida     = DOWNLOAD_DIR / f"{cfg['prefixo']}_Detalhado_{timestamp}.xlsx"
    wb.save(saida)
    
    log(f"  ✓ Excel geral salvo: {saida.name}")
    log(f"  ✓ Arquivos individuais de importação gerados na pasta: {pasta_importacao.name}")
    
    return pasta_importacao


# ══════════════════════════════════════════════════════════════════════════════
# ORQUESTRADOR ASYNC
# ══════════════════════════════════════════════════════════════════════════════

async def run_agent(card: str, modo: int, usar_cache: bool, detalhes_cache: dict, email: str = "", senha: str = ""):
    cfg = CARDS[card]
    log("=" * 55)
    log(f"  AGENTE DATAFRETE — {cfg['label']}")
    log(f"  Modo: {'GERAL' if modo == 1 else 'DETALHADO'}")
    log("=" * 55)

    DOWNLOAD_DIR.mkdir(exist_ok=True)

    if usar_cache:
        log("  ℹ  Modo cache — pulando extração no Datafrete")
        arquivo_bruto = _bruto_mais_recente(card)
        if arquivo_bruto is None:
            log("  ⚠ Nenhum bruto encontrado. Execute sem cache primeiro.")
            return
        detalhes = detalhes_cache
    else:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=HEADLESS, args=["--start-maximized"])
            context = await browser.new_context(
                accept_downloads=True, viewport={"width": 1440, "height": 900}
            )
            page = await context.new_page()
            try:
                await step_login(page, email, senha)
                await step_abrir_modal(page, cfg)
                await step_selecionar_visualizacao(page, cfg)
                arquivo_bruto = await step_exportar_resumo(page, cfg)
                detalhes = {}
                if modo == 2:
                    detalhes = await step_drill_down(page, context, card, cfg)
            except PlaywrightTimeout as e:
                log(f"  ❌ Timeout: {e}")
                try: await page.screenshot(path="erro_timeout.png")
                except: pass
                raise
            except Exception as e:
                log(f"  ❌ Erro: {e}")
                try: await page.screenshot(path="erro_geral.png")
                except: pass
                raise
            finally:
                await browser.close()

    linhas, total_nf, total_ct, por_unidade = ler_arquivo_bruto(arquivo_bruto)

    if modo == 1:
        relatorio = gerar_excel_geral(linhas, total_nf, total_ct, por_unidade, cfg, card)
    else:
        relatorio = gerar_excel_detalhado(
            linhas, total_nf, total_ct, por_unidade, detalhes, cfg, card
        )

    log("=" * 55)
    log("  ✅ CONCLUÍDO")
    log(f"  📄 Bruto:     {arquivo_bruto.name}")
    log(f"  📊 Relatório: {relatorio.name}")
    log("=" * 55)

    abrir_arquivo(relatorio)


# ══════════════════════════════════════════════════════════════════════════════
# INTERFACE GRÁFICA (GUI)
# ══════════════════════════════════════════════════════════════════════════════

class DatafreteApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("FOTUS — Agente de Automação Datafrete")
        self.geometry("750x700")
        ctk.set_appearance_mode("System")
        
        # Conecta a função de log global com a nossa UI
        global UI_LOG_CALLBACK
        UI_LOG_CALLBACK = self.append_log
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1) # Faz o log expandir

        # ── Título ──
        self.label_titulo = ctk.CTkLabel(self, text="Agente de Relatórios Datafrete", font=ctk.CTkFont(size=26, weight="bold"))
        self.label_titulo.grid(row=0, column=0, padx=20, pady=(25, 15))

        # ── Frame de Opções ──
        self.frame_opcoes = ctk.CTkFrame(self)
        self.frame_opcoes.grid(row=1, column=0, padx=40, pady=10, sticky="nsew")
        self.frame_opcoes.grid_columnconfigure(0, weight=1)

        # 1. Relatório
        self.label_card = ctk.CTkLabel(self.frame_opcoes, text="1. Selecione o Relatório:", font=ctk.CTkFont(weight="bold", size=14))
        self.label_card.grid(row=0, column=0, pady=(20, 5))
        
        self.var_card = ctk.StringVar(value="A")
        self.frame_radios = ctk.CTkFrame(self.frame_opcoes, fg_color="transparent")
        self.frame_radios.grid(row=1, column=0, pady=5)
        self.rb_arquivos = ctk.CTkRadioButton(self.frame_radios, text="Pendências de Arquivos", variable=self.var_card, value="A")
        self.rb_arquivos.grid(row=0, column=0, padx=15)
        self.rb_entrega = ctk.CTkRadioButton(self.frame_radios, text="Pendências de Entrega", variable=self.var_card, value="B")
        self.rb_entrega.grid(row=0, column=1, padx=15)

        # 2. Modo
        self.label_modo = ctk.CTkLabel(self.frame_opcoes, text="2. Selecione o Modo:", font=ctk.CTkFont(weight="bold", size=14))
        self.label_modo.grid(row=2, column=0, pady=(15, 5))
        
        self.menu_modo = ctk.CTkOptionMenu(self.frame_opcoes, values=["GERAL / MACRO", "DETALHADO / ANALÍTICO"], width=280)
        self.menu_modo.grid(row=3, column=0, pady=5)

        # 3. Credenciais de Acesso
        self.label_credenciais = ctk.CTkLabel(self.frame_opcoes, text="3. Credenciais do Datafrete:", font=ctk.CTkFont(weight="bold", size=14))
        self.label_credenciais.grid(row=4, column=0, pady=(15, 5))

        self.frame_login = ctk.CTkFrame(self.frame_opcoes, fg_color="transparent")
        self.frame_login.grid(row=5, column=0, pady=5)
        
        self.entry_email = ctk.CTkEntry(self.frame_login, placeholder_text="E-mail", width=250)
        self.entry_email.grid(row=0, column=0, padx=10)
        
        self.entry_senha = ctk.CTkEntry(self.frame_login, placeholder_text="Senha", show="*", width=200)
        self.entry_senha.grid(row=0, column=1, padx=10)

        # 4. Cache
        self.var_cache = ctk.IntVar(value=0)
        self.cb_cache = ctk.CTkCheckBox(
            self.frame_opcoes, 
            text="REPROCESSAR planilhas já extraídas (Não acessar o Datafrete)", 
            variable=self.var_cache,
            font=ctk.CTkFont(weight="bold")
        )
        self.cb_cache.grid(row=6, column=0, pady=(20, 25))

        # ── Botão Iniciar ──
        self.btn_run = ctk.CTkButton(self, text="INICIAR AUTOMAÇÃO", command=self.start_automation_thread, 
                                     fg_color="#1B4F8A", hover_color="#153e6d", font=ctk.CTkFont(weight="bold", size=15), height=45, width=220)
        self.btn_run.grid(row=2, column=0, pady=20)

        # ── Área de Logs ──
        self.label_log = ctk.CTkLabel(self, text="Log de Execução:", font=ctk.CTkFont(weight="bold"))
        self.label_log.grid(row=3, column=0, padx=40, pady=(5, 0), sticky="w")
        
        self.text_log = ctk.CTkTextbox(self, height=180, font=ctk.CTkFont(family="Consolas", size=12))
        self.text_log.grid(row=4, column=0, padx=40, pady=(0, 25), sticky="nsew")

    def append_log(self, msg):
        """Método seguro para atualizar a interface a partir de outra thread."""
        self.after(0, self._insert_text, msg)

    def _insert_text(self, msg):
        self.text_log.insert("end", msg + "\n")
        self.text_log.see("end")

    def start_automation_thread(self):
        """Lança a automação em uma thread separada para não travar a UI"""
        self.btn_run.configure(state="disabled")
        self.text_log.delete("1.0", "end")
        
        escolha_card = "arquivos" if self.var_card.get() == "A" else "entrega"
        escolha_modo = 1 if self.menu_modo.get() == "GERAL / MACRO" else 2
        usar_cache_ui = bool(self.var_cache.get())
        
        email = self.entry_email.get().strip()
        senha = self.entry_senha.get().strip()
        
        if not usar_cache_ui and (not email or not senha):
            self.append_log("⚠ ERRO: Preencha o E-mail e a Senha para rodar a automação (ou marque o cache).")
            self.btn_run.configure(state="normal")
            return
            
        detalhes_cache = {}
        usar_cache_final = False

        if usar_cache_ui:
            cache_data, tem_cache = verificar_cache(escolha_card)
            if tem_cache:
                detalhes_cache = cache_data
                usar_cache_final = True
            else:
                log("⚠ Nenhum cache válido encontrado. Extração no Datafrete será iniciada.")
                if not email or not senha:
                    self.append_log("⚠ ERRO: Preencha E-mail e Senha, pois não há cache disponível para pular o login.")
                    self.btn_run.configure(state="normal")
                    return
        
        thread = threading.Thread(target=self.run_async_wrapper, args=(escolha_card, escolha_modo, usar_cache_final, detalhes_cache, email, senha))
        thread.start()

    def run_async_wrapper(self, card, modo, usar_cache, detalhes_cache, email, senha):
        """Wrapper para rodar o loop de eventos async dentro da thread secundária"""
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(run_agent(card, modo, usar_cache, detalhes_cache, email, senha))
        except Exception as e:
            log(f"ERRO CRÍTICO NA EXECUÇÃO: {e}")
        finally:
            loop.close()
            # Reativa o botão na thread principal
            self.after(0, lambda: self.btn_run.configure(state="normal"))


if __name__ == "__main__":
    app = DatafreteApp()
    app.mainloop()